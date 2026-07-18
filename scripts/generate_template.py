"""Generate the Korea-Uzbekistan business tracker Excel workbook.

Run with: python3 scripts/generate_template.py

This script refuses to overwrite the output file if it already exists,
so it is safe to re-run without risk of destroying collected data.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "korea_uzbekistan_companies.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COMPANY_COLUMNS = [
    ("Company Name (EN)", 26),
    ("Company Name (KR)", 20),
    ("Uzbekistan Entity Name", 26),
    ("Sector", 22),
    ("Sub-sector / Product", 26),
    ("Parent Company / Korea HQ", 24),
    ("HQ Established Year", 14),
    ("Uzbekistan Entry Year", 14),
    ("Investment Amount (USD)", 18),
    ("Investment Type", 16),
    ("Annual Revenue (most recent)", 20),
    ("Revenue Year", 12),
    ("Net Profit (most recent)", 18),
    ("Profit Year", 12),
    ("Employees (Global)", 14),
    ("Employees (Uzbekistan)", 16),
    ("Project Status", 16),
    ("Uzbekistan City / Location", 18),
    ("Project Description", 40),
    ("Korea HQ Website", 24),
    ("Uzbekistan Website / Careers Page", 28),
    ("Source Type", 18),
    ("Source URL(s)", 40),
    ("Date Added", 12),
    ("Last Updated", 12),
    ("Notes", 30),
]

DATA_DICTIONARY = [
    ("Company Name (EN)", "Official company name in English."),
    ("Company Name (KR)", "Official company name in Korean (as it appears in Korean press/registries)."),
    ("Uzbekistan Entity Name", "Name of the local subsidiary, branch, JV, or project name in Uzbekistan, if different from the parent."),
    ("Sector", "Primary industry. Choose from the dropdown; add new sectors to the list in this script if needed."),
    ("Sub-sector / Product", "More specific product/service line, e.g. 'auto parts - wiring harness', 'CT scanners'."),
    ("Parent Company / Korea HQ", "The ultimate Korean parent company, if the Uzbekistan entity is a subsidiary or JV partner."),
    ("HQ Established Year", "Year the Korean parent company was founded."),
    ("Uzbekistan Entry Year", "Year the company entered the Uzbekistan market or announced the project."),
    ("Investment Amount (USD)", "Disclosed or estimated investment amount in USD. Note the source and whether it's committed vs. actual."),
    ("Investment Type", "Greenfield, Joint Venture, M&A, Loan/Financing, Grant/Aid, Distribution/Branch, or Unknown."),
    ("Annual Revenue (most recent)", "Most recently reported annual revenue of the parent company (or local entity, if disclosed). Note currency in Notes if not USD."),
    ("Revenue Year", "Fiscal year the revenue figure corresponds to."),
    ("Net Profit (most recent)", "Most recently reported net profit of the parent company (or local entity, if disclosed)."),
    ("Profit Year", "Fiscal year the profit figure corresponds to."),
    ("Employees (Global)", "Total employee count of the parent company, company-wide."),
    ("Employees (Uzbekistan)", "Employee count at the Uzbekistan operation specifically, if disclosed."),
    ("Project Status", "Planned, Under Construction, Operational, Completed, Withdrawn, or Unknown."),
    ("Uzbekistan City / Location", "City or region of the Uzbekistan operation (e.g. Tashkent, Andijan, Navoi)."),
    ("Project Description", "Short free-text summary of what the company/project does in Uzbekistan."),
    ("Korea HQ Website", "Official website of the Korean parent company."),
    ("Uzbekistan Website / Careers Page", "Local site or careers page, if one exists -- useful for your own job search."),
    ("Source Type", "KOTRA Report, EXIM Bank FDI Stats, DART Disclosure, Korean News, Uzbek News, Company Website, LinkedIn/Job Posting, or Other."),
    ("Source URL(s)", "Link(s) to where you found this information. Separate multiple links with a semicolon."),
    ("Date Added", "Date (YYYY-MM-DD) you first added this row."),
    ("Last Updated", "Date (YYYY-MM-DD) you last verified or updated this row."),
    ("Notes", "Anything else worth remembering: currency caveats, conflicting sources, follow-up needed, etc."),
]

SOURCES_COLUMNS = ["Date Checked", "Source Name", "URL", "Sector / Topic", "Notes"]
SEARCH_LOG_COLUMNS = ["Date", "Search Engine", "Query (EN/KR)", "Sector", "Companies Found", "Notes"]

SECTOR_OPTIONS = "Automotive,Semiconductor/Electronics,Healthcare/Pharma,Energy/Oil & Gas,Textile,Construction/Infrastructure,IT/Software,Banking/Finance,Agriculture,Logistics,Education,Other"
INVESTMENT_TYPE_OPTIONS = "Greenfield,Joint Venture,M&A,Loan/Financing,Grant/Aid,Distribution/Branch,Unknown"
STATUS_OPTIONS = "Planned,Under Construction,Operational,Completed,Withdrawn,Unknown"
SOURCE_TYPE_OPTIONS = "KOTRA Report,EXIM Bank FDI Stats,DART Disclosure,Korean News,Uzbek News,Company Website,LinkedIn/Job Posting,Other"


def style_header(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    ws.freeze_panes = "A2"


def add_dropdown(ws, options_csv, col_letter, max_row=500):
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def build():
    if os.path.exists(OUTPUT_PATH):
        raise SystemExit(f"Refusing to overwrite existing file: {OUTPUT_PATH}")

    wb = Workbook()

    # Companies sheet
    ws = wb.active
    assert ws is not None
    ws.title = "Companies"
    for idx, (name, width) in enumerate(COMPANY_COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=name)
        ws.column_dimensions[get_column_letter(idx)].width = width
    style_header(ws, len(COMPANY_COLUMNS))

    col_index = {name: i + 1 for i, (name, _) in enumerate(COMPANY_COLUMNS)}
    add_dropdown(ws, SECTOR_OPTIONS, get_column_letter(col_index["Sector"]))
    add_dropdown(ws, INVESTMENT_TYPE_OPTIONS, get_column_letter(col_index["Investment Type"]))
    add_dropdown(ws, STATUS_OPTIONS, get_column_letter(col_index["Project Status"]))
    add_dropdown(ws, SOURCE_TYPE_OPTIONS, get_column_letter(col_index["Source Type"]))

    # Data Dictionary sheet
    dd = wb.create_sheet("Data Dictionary")
    dd.append(["Column Name", "Description"])
    for row in DATA_DICTIONARY:
        dd.append(list(row))
    style_header(dd, 2)
    dd.column_dimensions["A"].width = 30
    dd.column_dimensions["B"].width = 90
    for r in range(2, dd.max_row + 1):
        dd.cell(row=r, column=2).alignment = WRAP

    # Sources sheet
    src = wb.create_sheet("Sources")
    src.append(SOURCES_COLUMNS)
    style_header(src, len(SOURCES_COLUMNS))
    widths = [14, 26, 50, 20, 40]
    for i, w in enumerate(widths, start=1):
        src.column_dimensions[get_column_letter(i)].width = w

    # Search Log sheet
    log = wb.create_sheet("Search Log")
    log.append(SEARCH_LOG_COLUMNS)
    style_header(log, len(SEARCH_LOG_COLUMNS))
    widths = [12, 16, 40, 20, 16, 40]
    for i, w in enumerate(widths, start=1):
        log.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
